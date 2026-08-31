#!/usr/bin/env python3
"""Export FOOD exact-span per-class results for SPSR-Net and external baselines."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parents[1] / "runs" / "food_spsr_and_baselines_per_class.xlsx"

# Every tuple is: (precision, recall, f1, gold).  Values are percentages.
RESULTS = {
    "SPSR-Net": {
        "FC": (98.50, 100.00, 99.24, 197), "FN": (96.43, 72.97, 83.08, 37),
        "IN": (98.03, 95.51, 96.75, 156), "IV": (99.32, 97.35, 98.33, 151),
        "ORG": (79.41, 77.14, 78.26, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (42.86, 50.00, 46.15, 6), "SNum": (96.33, 98.74, 97.52, 239),
        "Overall": (96.56, 95.62, 96.09, 822),
    },
    "CNN_Nested_NER": {
        "FC": (100.00, 100.00, 100.00, 197), "FN": (97.22, 94.59, 95.89, 37),
        "IN": (99.33, 95.51, 97.39, 156), "IV": (100.00, 97.35, 98.66, 151),
        "ORG": (87.50, 80.00, 83.58, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (75.00, 50.00, 60.00, 6), "SNum": (97.55, 100.00, 98.76, 239),
        "Overall": (98.40, 97.08, 97.73, 822),
    },
    "LEBERT*": {
        "FC": (99.49, 99.49, 99.49, 196), "FN": (100.00, 100.00, 100.00, 36),
        "IN": (99.33, 95.48, 97.37, 155), "IV": (100.00, 97.35, 98.66, 151),
        "ORG": (86.11, 88.57, 87.32, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (30.77, 66.67, 42.11, 6), "SNum": (97.51, 98.33, 97.92, 239),
        "Overall": (97.31, 97.19, 97.25, 819),
    },
    "W2NER": {
        "FC": (100.00, 100.00, 100.00, 197), "FN": (100.00, 29.73, 45.83, 37),
        "IN": (99.33, 95.51, 97.39, 156), "IV": (100.00, 97.35, 98.66, 151),
        "ORG": (75.00, 77.14, 76.06, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (100.00, 33.33, 50.00, 6), "SNum": (99.16, 99.16, 99.16, 239),
        "Overall": (98.47, 93.67, 96.01, 822),
    },
    "locate-and-label": {
        "FC": (98.47, 97.97, 98.22, 197), "FN": (100.00, 75.68, 86.15, 37),
        "IN": (97.32, 92.95, 95.08, 156), "IV": (99.26, 89.40, 94.08, 151),
        "ORG": (82.05, 91.43, 86.49, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (50.00, 33.33, 40.00, 6), "SNum": (97.15, 100.00, 98.56, 239),
        "Overall": (96.99, 94.16, 95.56, 822),
    },
    "DiffusionNER": {
        "FC": (94.47, 95.43, 94.95, 197), "FN": (100.00, 2.70, 5.26, 37),
        "IN": (100.00, 94.87, 97.37, 156), "IV": (100.00, 96.69, 98.32, 151),
        "ORG": (83.33, 85.71, 84.51, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (100.00, 16.67, 28.57, 6), "SNum": (99.51, 85.77, 92.13, 239),
        "Overall": (97.56, 87.47, 92.24, 822),
    },
    "Flat-Lattice-Transformer†": {
        "FC": (98.50, 100.00, 99.24, 197), "FN": (91.18, 83.78, 87.32, 37),
        "IN": (38.68, 78.85, 51.90, 156), "IV": (69.35, 91.39, 78.86, 151),
        "ORG": (100.00, 57.14, 72.73, 35), "PER": (0.00, 0.00, 0.00, 1),
        "SN": (0.00, 0.00, 0.00, 6), "SNum": (86.09, 82.85, 84.43, 239),
        "Overall": (70.49, 86.01, 77.48, 822),
    },
}

CLASSES = ["FC", "FN", "IN", "IV", "ORG", "PER", "SN", "SNum"]
MODELS = list(RESULTS)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def fit_columns(ws):
    for column in ws.columns:
        index = column[0].column
        width = min(34, max(len(str(cell.value or "")) for cell in column) + 2)
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    note = wb.active
    note.title = "说明"
    note["A1"] = "FOOD test：SPSR-Net 与对比模型逐类别结果"
    note["A1"].font = TITLE_FONT
    note["A3"] = "统一口径"
    note["B3"] = "实体边界与实体类型均完全一致；数值均为百分比。"
    note["A4"] = "类别样本数"
    note["B4"] = "FC=197, FN=37, IN=156, IV=151, ORG=35, PER=1, SN=6, SNum=239；原始 test 共 822 个实体。"
    note["A5"] = "LEBERT*"
    note["B5"] = "平面 BIO 模型，3 个重叠实体不可表达，评测 gold=819（FC/FN/IN 分别少 1）。"
    note["A6"] = "Flat-Lattice-Transformer†"
    note["B6"] = "平面 BMES 模型；此处是对原始 822 个实体的计分，嵌套实体会使其处于结构性劣势。"
    note.column_dimensions["A"].width = 28
    note.column_dimensions["B"].width = 110

    overall = wb.create_sheet("总体")
    overall.append(["模型", "Precision", "Recall", "F1", "Gold"])
    for model in MODELS:
        p, r, f1, gold = RESULTS[model]["Overall"]
        overall.append([model, p, r, f1, gold])
    style_header(overall)
    for row in overall.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "0.00"
    fit_columns(overall)

    detail = wb.create_sheet("逐类明细")
    detail.append(["模型", "类别", "Gold", "Precision", "Recall", "F1"])
    for model in MODELS:
        for label in CLASSES:
            p, r, f1, gold = RESULTS[model][label]
            detail.append([model, label, gold, p, r, f1])
    style_header(detail)
    for row in detail.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = "0.00"
    detail.auto_filter.ref = detail.dimensions
    fit_columns(detail)

    f1_sheet = wb.create_sheet("类别F1横向")
    f1_sheet.append(["模型"] + [f"{label} (Gold={RESULTS['SPSR-Net'][label][3]})" for label in CLASSES])
    for model in MODELS:
        f1_sheet.append([model] + [RESULTS[model][label][2] for label in CLASSES])
    style_header(f1_sheet)
    for row in f1_sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.00"
    fit_columns(f1_sheet)

    prf = wb.create_sheet("类别PRF横向")
    prf.append(["模型"] + [f"{label} P/R/F1" for label in CLASSES])
    for model in MODELS:
        prf.append([model] + ["{:.2f}/{:.2f}/{:.2f}".format(*RESULTS[model][label][:3]) for label in CLASSES])
    style_header(prf)
    fit_columns(prf)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
